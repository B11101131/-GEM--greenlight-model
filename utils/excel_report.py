"""
Excel 報告產生器

使用 openpyxl 產生格式化的 Excel 報告，包含摘要、模擬數據、
損益分析、現金流、敏感度和 VaR 等多個工作表。
"""
import io
import logging
from typing import Dict, Optional
from datetime import datetime

import numpy as np
import pandas as pd

logger = logging.getLogger(__name__)


def generate_excel_report(
    sim_result: Dict,
    sensitivity_df: Optional[pd.DataFrame] = None,
    cash_flow_df: Optional[pd.DataFrame] = None,
    var_results: Optional[list] = None,
    breakeven_result: Optional[object] = None,
    fig_profit: Optional[object] = None,
    fig_tornado: Optional[object] = None,
) -> bytes:
    """
    產生完整的 Excel 報告。

    NOTE: 使用 openpyxl 引擎。若未安裝則回報錯誤。
          圖表以 PNG 圖片嵌入 Excel。

    Args:
        sim_result: 模擬結果字典（含 params 和 results）
        sensitivity_df: 敏感度分析 DataFrame
        cash_flow_df: 現金流 DataFrame
        var_results: VaR 計算結果列表
        breakeven_result: 損益兩平結果物件
        fig_profit: 獲利分佈 matplotlib Figure
        fig_tornado: 龍捲風圖 matplotlib Figure

    Returns:
        Excel 檔案的 bytes
    """
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils.dataframe import dataframe_to_rows
    except ImportError:
        raise ImportError("需要安裝 openpyxl 套件：pip install openpyxl")

    wb = Workbook()
    now = datetime.now().strftime('%Y-%m-%d %H:%M')

    # 共用樣式
    header_font = Font(name='Calibri', size=12, bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='2C3E50', end_color='2C3E50', fill_type='solid')
    title_font = Font(name='Calibri', size=14, bold=True)
    data_font = Font(name='Calibri', size=11)
    header_alignment = Alignment(horizontal='center', vertical='center')
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin'),
    )

    def _style_header_row(ws, row_num: int, max_col: int) -> None:
        """套用標題列樣式"""
        for col in range(1, max_col + 1):
            cell = ws.cell(row=row_num, column=col)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border

    def _auto_width(ws) -> None:
        """自動調整欄寬"""
        for col in ws.columns:
            max_length = 0
            col_letter = col[0].column_letter
            for cell in col:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[col_letter].width = min(max_length + 4, 30)

    # ============================================================
    # 工作表 1：摘要
    # ============================================================
    ws_summary = wb.active
    ws_summary.title = "摘要"

    ws_summary.cell(row=1, column=1, value="🎬 影視專案綠燈評估報告").font = Font(
        name='Calibri', size=16, bold=True
    )
    ws_summary.cell(row=2, column=1, value=f"產生時間：{now}").font = data_font

    # 專案資訊
    project_info = [
        ("專案名稱", sim_result.get('project_name', 'N/A')),
        ("類型", sim_result.get('genre', 'N/A')),
        ("製作預算 (百萬 TWD)", f"{sim_result.get('budget', 0):.1f}"),
        ("行銷宣發費 (百萬 TWD)", f"{sim_result.get('marketing_pa', 0):.1f}"),
        ("總成本 (百萬 TWD)", f"{sim_result.get('budget', 0) + sim_result.get('marketing_pa', 0):.1f}"),
    ]

    row = 4
    ws_summary.cell(row=row, column=1, value="項目").font = header_font
    ws_summary.cell(row=row, column=1).fill = header_fill
    ws_summary.cell(row=row, column=2, value="數值").font = header_font
    ws_summary.cell(row=row, column=2).fill = header_fill
    _style_header_row(ws_summary, row, 2)

    for label, value in project_info:
        row += 1
        ws_summary.cell(row=row, column=1, value=label).font = data_font
        ws_summary.cell(row=row, column=2, value=value).font = data_font

    # 關鍵指標
    row += 2
    ws_summary.cell(row=row, column=1, value="關鍵指標").font = title_font

    metrics = [
        ("預期淨利 (百萬 TWD)", f"{sim_result.get('expected_profit', 0):.1f}"),
        ("預期投資報酬率 (%)", f"{sim_result.get('expected_roi', 0):.1f}"),
        ("虧損機率 (%)", f"{sim_result.get('prob_loss', 0):.1f}"),
        ("模擬次數", f"{sim_result.get('simulation_count', 0):,}"),
    ]

    row += 1
    ws_summary.cell(row=row, column=1, value="指標").font = header_font
    ws_summary.cell(row=row, column=1).fill = header_fill
    ws_summary.cell(row=row, column=2, value="數值").font = header_font
    ws_summary.cell(row=row, column=2).fill = header_fill
    _style_header_row(ws_summary, row, 2)

    for label, value in metrics:
        row += 1
        ws_summary.cell(row=row, column=1, value=label).font = data_font
        ws_summary.cell(row=row, column=2, value=value).font = data_font

    _auto_width(ws_summary)

    # ============================================================
    # 工作表 2：模擬數據統計
    # ============================================================
    ws_sim = wb.create_sheet("模擬數據")

    sim_stats = [
        ("統計量", "數值"),
        ("平均淨利", f"{sim_result.get('expected_profit', 0):.2f}"),
        ("標準差", f"{np.std(sim_result.get('net_profit', [0])):.2f}" if 'net_profit' in sim_result else "N/A"),
        ("P5 (悲觀)", f"{sim_result.get('p5', 0):.2f}"),
        ("P25", f"{np.percentile(sim_result.get('net_profit', [0]), 25):.2f}" if 'net_profit' in sim_result else "N/A"),
        ("中位數", f"{np.median(sim_result.get('net_profit', [0])):.2f}" if 'net_profit' in sim_result else "N/A"),
        ("P75", f"{np.percentile(sim_result.get('net_profit', [0]), 75):.2f}" if 'net_profit' in sim_result else "N/A"),
        ("P95 (樂觀)", f"{sim_result.get('p95', 0):.2f}"),
    ]

    for i, (label, value) in enumerate(sim_stats, 1):
        ws_sim.cell(row=i, column=1, value=label).font = data_font if i > 1 else header_font
        ws_sim.cell(row=i, column=2, value=value).font = data_font if i > 1 else header_font
        if i == 1:
            _style_header_row(ws_sim, i, 2)

    _auto_width(ws_sim)

    # ============================================================
    # 工作表 3：損益兩平
    # ============================================================
    if breakeven_result:
        ws_be = wb.create_sheet("損益兩平")
        be_data = [
            ("項目", "數值"),
            ("損益兩平票房 (百萬)", f"{breakeven_result.breakeven_box_office:.1f}"),
            ("損益兩平收入 (百萬)", f"{breakeven_result.breakeven_revenue:.1f}"),
            ("達成機率 (%)", f"{breakeven_result.prob_breakeven:.1f}"),
            ("安全邊際", f"{breakeven_result.safety_margin:.1%}"),
        ]

        for i, (label, value) in enumerate(be_data, 1):
            ws_be.cell(row=i, column=1, value=label).font = data_font if i > 1 else header_font
            ws_be.cell(row=i, column=2, value=value).font = data_font if i > 1 else header_font
            if i == 1:
                _style_header_row(ws_be, i, 2)

        _auto_width(ws_be)

    # ============================================================
    # 工作表 4：現金流
    # ============================================================
    if cash_flow_df is not None and not cash_flow_df.empty:
        ws_cf = wb.create_sheet("現金流")

        for r_idx, row_data in enumerate(dataframe_to_rows(cash_flow_df, index=False, header=True), 1):
            for c_idx, value in enumerate(row_data, 1):
                cell = ws_cf.cell(row=r_idx, column=c_idx, value=value)
                cell.font = data_font
                if r_idx == 1:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment

        _auto_width(ws_cf)

    # ============================================================
    # 工作表 5：敏感度分析
    # ============================================================
    if sensitivity_df is not None and not sensitivity_df.empty:
        ws_sens = wb.create_sheet("敏感度分析")

        display_cols = ['Parameter', 'Low Impact', 'High Impact', 'Impact Range']
        sens_display = sensitivity_df[display_cols].copy()
        sens_display = sens_display.sort_values('Impact Range', ascending=False)

        for r_idx, row_data in enumerate(dataframe_to_rows(sens_display, index=False, header=True), 1):
            for c_idx, value in enumerate(row_data, 1):
                cell = ws_sens.cell(row=r_idx, column=c_idx)
                if isinstance(value, float):
                    cell.value = round(value, 2)
                else:
                    cell.value = value
                cell.font = data_font
                if r_idx == 1:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = header_alignment

        _auto_width(ws_sens)

    # ============================================================
    # 工作表 6：VaR 風險值
    # ============================================================
    if var_results:
        ws_var = wb.create_sheet("VaR")

        var_headers = ["信心水準 (%)", "VaR (百萬)", "CVaR (百萬)", "最大損失 (百萬)", "解讀"]
        for c_idx, header in enumerate(var_headers, 1):
            ws_var.cell(row=1, column=c_idx, value=header)
        _style_header_row(ws_var, 1, len(var_headers))

        for r_idx, vr in enumerate(var_results, 2):
            ws_var.cell(row=r_idx, column=1, value=f"{vr.confidence_level:.0f}").font = data_font
            ws_var.cell(row=r_idx, column=2, value=f"{vr.var_value:.2f}").font = data_font
            ws_var.cell(row=r_idx, column=3, value=f"{vr.cvar_value:.2f}").font = data_font
            ws_var.cell(row=r_idx, column=4, value=f"{vr.max_loss:.2f}").font = data_font
            ws_var.cell(row=r_idx, column=5, value=vr.interpretation).font = data_font

        _auto_width(ws_var)

    # ============================================================
    # 嵌入圖表圖片
    # ============================================================
    try:
        from openpyxl.drawing.image import Image as XlImage

        if fig_profit is not None:
            img_buf = io.BytesIO()
            fig_profit.savefig(img_buf, format='png', dpi=150, bbox_inches='tight')
            img_buf.seek(0)
            img = XlImage(img_buf)
            img.width = 600
            img.height = 400
            ws_summary.add_image(img, f"D4")

        if fig_tornado is not None:
            img_buf = io.BytesIO()
            fig_tornado.savefig(img_buf, format='png', dpi=150, bbox_inches='tight')
            img_buf.seek(0)
            img = XlImage(img_buf)
            img.width = 600
            img.height = 400

            # HACK: 放在敏感度工作表中。若該工作表不存在，放在摘要頁
            target_ws = wb["敏感度分析"] if "敏感度分析" in wb.sheetnames else ws_summary
            target_ws.add_image(img, "D1")

    except Exception as e:
        logger.warning(f"無法嵌入圖表圖片: {e}")

    # 輸出為 bytes
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output.getvalue()
